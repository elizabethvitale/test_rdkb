import subprocess
import random
import socket
import fcntl
import struct
import time
from openpyxl import Workbook

#gets a random hex value with a valid raspberry pi address
def get_hex():
    return [ 0xb8, 0x27, 0xeb, random.randint(0x00, 0x7f), random.randint(0x00, 0xff), random.randint(0x00, 0xff) ]
    
#turns a random hex value into a fully-mapped MAC address
def to_string(address):
    return ':'.join(map(lambda x: "%02x" % x, address))

#changes the MAC address to a randomly generated hex value MAC address
def run_times(x):
    for i in range(x):
        mac = to_string(get_hex())
        configure_ip(mac, i, 0)
            
#fixes case where wifi fails connection the first time            
def configure_ip(mac, i, j):
    try:
        subprocess.call(["sudo", "ip", "link", "set", "dev", "wlan0", "down"]) 
        subprocess.call(["sudo", "ip", "link", "set", "dev", "wlan0", "address", mac])
        subprocess.call(["sudo", "ip", "link", "set", "dev", "wlan0", "up"])
        subprocess.call(["sudo", "ip", "link"])
        print(i)
        time.sleep(15)
        
        ip = get_ip('wlan0')
        ip_last_digits = int(ip[10:])
        
        ws.cell(row = i+2, column = 1).value = ip
        ws.cell(row = i+2, column = 2).value = ip_last_digits
        ws.cell(row = i+2, column = 3).value = mac
    except Exception as e:
        j = j + 1
        if(j < 3):
            configure_ip(mac, i, j)
        elif(j >= 3):
            ip = get_ip('wlan0')
            ws.cell(row = i+2, column = 1).value = ip
            ws.cell(row = i+2, column = 3).value = mac

            
#gets the IPV4 address
def get_ip(ifname):
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    return socket.inet_ntoa(fcntl.ioctl(
        s.fileno(),
        0x8915,  # SIOCGIFADDR
        struct.pack('256s',bytes( ifname[:15],'utf-8'))
    )[20:24])

    
#amount of times to generate a random MAC address        
subprocess.call(["sudo", "ip", "link", "set", "wlan0", "promisc", "on"])
wb = Workbook()
ws = wb.active
ws.title = "Test Data"
ws['A1'] = "Total IPv4"
ws['B1'] = "IPv4 Ending"
ws['C1'] = "MAC Address"

run_times(254)
wb.save('wifi_test_#1.xlsx')
